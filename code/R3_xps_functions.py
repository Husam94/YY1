import numpy as np
import pandas as pd

import os, glob, shutil 

#------------------------------------

import torch
import torch.nn as nn
from torch.optim import Adam, AdamW

#------------------------------------

import poseigen_seaside.basics as se
import poseigen_compass as co


from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font

import concurrent.futures



data_path = "../data/R3/"
os.makedirs(data_path, exist_ok=True)

os.chdir(data_path)

xpsfolder = se.NewFolder('xps_4')

#====================================================================

def Reset_DualOyster(dualoyster):
    #ds consists of a "conv" and a "dense" module. Need to go through each one, see if its a conv and reset if so. 


    for oyster in [dualoyster.OysterA, dualoyster.OysterB]: 
        lke, los = len(oyster.kE), len(oyster.O)

        for i in np.arange(lke): 
            if isinstance(oyster.kE[i], nn.Conv2d): 
                oyster.kE[i].reset_parameters()
        
        for i in np.arange(los): 
            if isinstance(oyster.O[i], nn.Conv2d): 
                oyster.O[i].reset_parameters()
        print('done reset mod')

    return dualoyster

#======================================

def XpsCopy(cur_pn, cur_xp_id, 
            past_xp_id, past_unit, 
            rewrite = False):
                         
    basetargo = cur_pn.replace(cur_xp_id, past_xp_id)[:-2] + str(past_unit) + '/'

    targnames = glob.glob(basetargo + '*')
    lbt = len(basetargo)

    targnames1 = [t for t in targnames if '.p' in t or '.pt' in t]
    newnames1 = [cur_pn + t[lbt:] for t in targnames1]

    print(targnames1, newnames1)


    if rewrite: 
        for n in newnames1: 
            if os.path.isfile(n): os.remove(n)
    for t,n in zip(targnames1, newnames1): 
        if os.path.isfile(n): break #means we already have it
        shutil.copyfile(t, n)

    targnames2 = [t for t in targnames if '.p' not in t and'.pt' not in t]
    newnames2 = [cur_pn + t[lbt:] for t in targnames2]
    if rewrite: 
        for n in newnames2: 
            if os.path.isdir(n): shutil.rmtree(n)
    for t, n in zip(targnames2, newnames2): 
        if os.path.isdir(n): break #means we already have it             
        shutil.copytree(t, n, dirs_exist_ok=True)
    
    print(f'finished copying for {cur_pn}')
    
    return


def XpsBootstrapEnsemb(combs, pathname, TCR_repeats, es_args, rewrite=False):
    iterx = 100

    def process_icom(args):
        icom, com = args
        pn_t = se.NewFolder(pathname + str(icom))
        pn_preds = [pn_t + '0_' + str(ir) + '_Preds.p' for ir in range(TCR_repeats)]
        pn_e = pn_t + 'boots_ensemb'
        if os.path.isfile(pn_e + '.p') and not rewrite:
            boots_ensemb = se.PickleLoad(pn_e)
        else:
            boots_ensemb = co.Bootstrapper(pn_preds, mode=[co.EnsembleScorer, es_args],
                                           iters=iterx, updates=20, return_idx=True)
            se.PickleDump(boots_ensemb, pn_t + 'boots_ensemb')
        print(f'---- Finished {icom} ----')
        return boots_ensemb[0], boots_ensemb[1]

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        results = list(executor.map(process_icom, combs))
    boots_scores, boots_idxs = zip(*results)
    return np.array(boots_scores), np.array(boots_idxs)

def XpsResults(pn_xp, xp_combs,
               TCR_repeats, es_args, allconfigs,
               icof_only = None,
               ref_ic = 0, rewrite = False): 
    
    pf_args = {'mode1': [se.RelativeChange, {'perc': True}]}

    bs_all = []
    bs_all_r2r = []

    for icof, (cof_name, cof) in enumerate(allconfigs.items()):

        if icof_only is not None: 
            if icof not in icof_only: continue 
                            
        pn_t1 = se.NewFolder(pn_xp + str(icof))

        boots_scores = XpsBootstrapEnsemb(xp_combs, pn_t1, TCR_repeats, es_args, rewrite = rewrite)[0] # Just get scores 

        bs_all.append(boots_scores)

        com = boots_scores[ref_ic]

        rel2refs = [se.PairwiseFuncer(boots_scores[ic], com, **pf_args) for ic, c in xp_combs]

        bs_all_r2r.append(rel2refs)

        print(f'**** Finished {icof} ****')
                    
    bs_all = np.stack(bs_all)
    se.PickleDump(bs_all, pn_xp + 'bs_all')

    bs_all_r2r = np.stack(bs_all_r2r)
    se.PickleDump(bs_all_r2r, pn_xp + 'bs_all_r2r' + '_' + str(ref_ic))

    return bs_all, bs_all_r2r

def XpsTables(pn_xp, xp_combs, xp_variables, ber2rs, allconfigs, 
              icof_only = None):
    
    onesided = None
    conf_alpha = 0.90

    print(ber2rs.shape)

    ber2rs = ber2rs[:, :, :, 2] ####################################

    ber2rs_mean = np.mean(ber2rs, axis = -1)
    ber2rs_se = np.std(ber2rs, axis = -1)
    ber2rs_low, ber2rs_high = co.BootstrapConfidenceInterval(ber2rs, alpha = conf_alpha, onesided=onesided, axis = -1)

    def sigo(low, high): 
        if onesided == 'lesser': sigi = 0 > high
        if onesided == 'greater': sigi = low > 0 
        if onesided == None: sigi = np.logical_or(0 > high, low > 0)
        return sigi
    
    ber2rs_sigo = sigo(ber2rs_low, ber2rs_high)

    #################################################

    per = pd.DataFrame(ber2rs_mean).T

    if len(xp_variables) > 1: 
        # per.index = pd.MultiIndex.from_tuples([tuple(x[1]) for x in cur_combs], names=cur_variables)
        per.index = [tuple(x[1]) for x in xp_combs]
        per.index.names = ['Combination']
    else: 
        per.index = [x[1][0] for x in xp_combs]
        per.index.names = xp_variables
    
    if len(xp_combs[0][1]) > 1: 
        multcolx = pd.MultiIndex.from_tuples([tuple(g[1]) for g in xp_combs])
    else: multcolx = [str(g[1][0]) for g in xp_combs]
    
    configos = list(allconfigs.keys())
    if icof_only is not None: configos = [cz for icz, cz in enumerate(configos) if icz in icof_only]

    per.columns = configos

    per_style = per.style.format(precision=1).background_gradient(axis = 1, vmin = -25, vmax = 0, cmap = 'Greys_r')


    per_style = per_style.apply(lambda _: 
                                np.where(ber2rs_sigo.reshape(-1, len(multcolx)).T, 
                                         'font-weight: bold', ''),
                                           axis = None)
        
    # #################################################

    ber2rs_stats = np.stack([ber2rs_mean, ber2rs_se, ber2rs_low, ber2rs_high, ber2rs_sigo], axis = -1)
    
    ber2rs_stats_tidy = [
        [co, *(c[:len(xp_variables)]), *ber2rs_stats[ico, ic]]
        for ico, co in enumerate(configos)
        for ic, c in xp_combs
    ]


    tups_data = [('Models', '', '')]

    tups_var = [*[('Unit', 'Subject',x) for x in xp_variables]]

    tups_stand = [('Stat', 'Mean', ''), ('Stat', 'SE', ''),
                  ('Stat', 'Percentile', '5th'), ('Stat', 'Percentile', '95th'),
                  ('Stat', 'Percentile', 'Sig.')]

    tups_all = [*tups_data, *tups_var, *tups_stand]

    pdx = pd.DataFrame(ber2rs_stats_tidy)
    multcol = pd.MultiIndex.from_tuples(tups_all)
    pdx.columns = multcol

    # Remove the 'ID' column from the DataFrame if it exists
    id_col = ('Unit', 'Subject', 'ID')
    if id_col in pdx.columns:
        pdx = pdx.drop(columns=[id_col])

    pdx[('Stat', 'Percentile', 'Sig.')] = pdx[('Stat', 'Percentile', 'Sig.')].astype('bool')

    def boldo(x): 
            if x[('Stat', 'Percentile', 'Sig.')] == True: 
                return ['font-weight: bold'] * len(x) 
            else: return [''] * len(x)

    variable_cols = [('Models', '', '')] + [('Unit', 'Subject', x) for x in xp_variables]
    pdx = pdx.sort_values(variable_cols)

    pdx_style = pdx.style.format(precision=4).apply(boldo, axis = 1).background_gradient(axis = 1, 
                                                                                         vmin = -25, vmax = 0, 
                                                                                         subset = [('Stat', 'Mean', '')], 
                                                                                         cmap = 'Greys_r')

    return per, per_style, pdx, pdx_style


def save_styled_tables(per_style, pdx_style, cur_pn, prefix_ext=None):

    table_styles = [
        {'selector': 'th', 'props': [('font-family', 'Times New Roman'), ('font-size', '11pt'), ('text-align', 'center')]},
        {'selector': 'td', 'props': [('font-family', 'Times New Roman'), ('font-size', '11pt'), ('text-align', 'center')]},
        {'selector': '.row_heading', 'props': [('font-family', 'Times New Roman'), ('font-size', '11pt'), ('text-align', 'center')]},
        {'selector': '.col_heading', 'props': [('font-family', 'Times New Roman'), ('font-size', '11pt'), ('text-align', 'center')]}
    ]
    per_style = per_style.set_table_styles(table_styles)
    pdx_style = pdx_style.set_table_styles(table_styles)

    tables_dir = se.NewFolder(xpsfolder + 'tables')

    prefix = os.path.basename(os.path.normpath(cur_pn))
    if prefix_ext is not None and prefix_ext != "":
        prefix = f"{prefix}_{prefix_ext}"

    per_file = os.path.join(tables_dir, f"{prefix}_tab_main_sty.xlsx")
    pdx_file = os.path.join(tables_dir, f"{prefix}_tab_extra_sty.xlsx")

    per_style.to_excel(per_file, engine='openpyxl')
    pdx_style.to_excel(pdx_file, engine='openpyxl')

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    font = Font(name="Times New Roman", size=11)

    def add_borders_align_font(filename, number_format=None, value_start_col=2):
        wb = load_workbook(filename)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows()):
            for j, cell in enumerate(row):
                orig_font = cell.font
                cell.font = Font(
                    name="Times New Roman",
                    size=11,
                    bold=orig_font.bold,
                    italic=orig_font.italic,
                    vertAlign=orig_font.vertAlign,
                    underline=orig_font.underline,
                    strike=orig_font.strike,
                    color=orig_font.color
                )
                cell.border = border
                cell.alignment = align
                # Set number format for value cells (skip header)
                if number_format and i > 0 and j >= value_start_col:
                    try:
                        float(cell.value)
                        cell.number_format = number_format
                    except (TypeError, ValueError):
                        pass
        wb.save(filename)

    # For your main table: 1 decimal
    add_borders_align_font(per_file, number_format="0.0", value_start_col=1)
    # For your extra table: 4 decimals (adjust value_start_col if needed)
    add_borders_align_font(pdx_file, number_format="0.0000", value_start_col=2)
    print(f"Saved styled tables as {per_file} and {pdx_file}")